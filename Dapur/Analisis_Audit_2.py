import datetime
import glob
import os
import re
import configparser
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.preprocessing import StandardScaler
from sklearn.cluster import DBSCAN, KMeans
from sklearn.ensemble import IsolationForest
from sklearn.neighbors import LocalOutlierFactor
from sklearn.svm import OneClassSVM
from sklearn.neural_network import MLPRegressor

print("--> Memulai pemrosesan data audit advanced machine learning (Analisis_Audit_2) ...")
config_file = "config.conf"
if not os.path.exists(config_file):
    with open(config_file, "w") as f:
        f.write("[FILTER]\nSHELL\nIRC\nZN\nGT\nFILTER\nJIMCO\nLAIN\nTOP 1\nOLI\n")

config = configparser.ConfigParser(allow_no_value=True)
config.read(config_file)
filter_produk = [str(key).upper() for key in config["FILTER"]]
if "LAIN" not in filter_produk:
    filter_produk.append("LAIN")

ptm_files = glob.glob("PTM*.xlsx")
if not ptm_files:
    print("--> File PTM*.xlsx tidak ditemukan.")
    exit()

ptm_file = ptm_files[0]
xl = pd.ExcelFile(ptm_file)

def sort_key_sheet(sheet_name):
    try:
        return datetime.datetime.strptime(sheet_name, "%m%y")
    except ValueError:
        return datetime.datetime.min

valid_sheets = [s for s in xl.sheet_names if re.match(r"^\d{4}$", s)]
valid_sheets.sort(key=sort_key_sheet)

if not valid_sheets:
    print("--> Tidak ada sheet bulanan (format MMYY) yang valid di file PTM.")
    exit()

id_months = {1:"Jan", 2:"Feb", 3:"Mar", 4:"Apr", 5:"Mei", 6:"Jun", 
             7:"Jul", 8:"Agu", 9:"Sep", 10:"Okt", 11:"Nov", 12:"Des"}

def clean_key(s):
    if s is None: return ""
    return re.sub(r"[^a-zA-Z0-9]", "", str(s)).lower()

font_title = Font(name="Calibri", size=14, bold=True, color="1F497D")
font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
fill_header = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")

fill_anomaly = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
thin_side = Side(border_style="thin", color="D9D9D9")
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
font_bold = Font(name="Calibri", size=11, bold=True)

monthly_macro_data = []
all_detailed_transactions = []

sisa_piutang_col = clean_key("Sisa Piutang")
nilai_faktur_col = clean_key("Nilai Faktur")
umur_japo_col = clean_key("Umur AR base on Tgl Faktur")
nama_penjual_col = clean_key("Nama Penjual")
nama_pelanggan_col = clean_key("Nama Pelanggan")
negara_pelanggan_col = clean_key("Negara Pelanggan")
no_faktur_col = clean_key("No. Faktur")

for sheet in valid_sheets:
    dt = datetime.datetime.strptime(sheet, "%m%y")
    col_label = f"{id_months[dt.month]}-{sheet[2:]}"
    
    df_all = pd.read_excel(ptm_file, sheet_name=sheet, header=None)
    header_idx = 0
    for idx, row in df_all.iterrows():
        row_clean = [clean_key(x) for x in row.dropna()]
        if sisa_piutang_col in row_clean or nama_penjual_col in row_clean:
            header_idx = idx
            break

    df = pd.read_excel(ptm_file, sheet_name=sheet, skiprows=header_idx)
    df_raw_for_loop = df.copy()
    df.columns = [clean_key(c) for c in df.columns]

    if not {sisa_piutang_col, umur_japo_col, nama_penjual_col, nama_pelanggan_col}.issubset(df.columns):
        continue

    df["days"] = df[umur_japo_col].astype(str).str.extract(r"(-?\d+)").astype(float).fillna(0)
    
    neg_p = df[negara_pelanggan_col].fillna("").astype(str).str.strip() if negara_pelanggan_col in df.columns else pd.Series([""]*len(df))
    nam_p = df[nama_pelanggan_col].fillna("").astype(str).str.strip()
    df["cust_id"] = np.where(neg_p.str.lower().isin(["", "nan", "none", "nat"]), nam_p, neg_p)

    df_raw_for_loop["Bulan_Sistem"] = col_label
    df_raw_for_loop["Cleaned_Sisa_Piutang"] = pd.to_numeric(df[sisa_piutang_col], errors="coerce").fillna(0)
    df_raw_for_loop["Cleaned_Nilai_Faktur"] = pd.to_numeric(df[nilai_faktur_col], errors="coerce").fillna(0) if nilai_faktur_col in df.columns else df_raw_for_loop["Cleaned_Sisa_Piutang"]
    df_raw_for_loop["Cleaned_Days"] = df["days"]
    df_raw_for_loop["Cleaned_Cust_ID"] = df["cust_id"]
    all_detailed_transactions.append(df_raw_for_loop)

df_micro_master = pd.concat(all_detailed_transactions, ignore_index=True)
final_output = "Laporan_Analisis_Prosedur_Audit.xlsx"
if os.path.exists(final_output):
    wb_new = openpyxl.load_workbook(final_output)
    print(f"--> Berhasil memuat berkas eksisting '{final_output}'. Menyisipkan prosedur baru...")
else:
    wb_new = openpyxl.Workbook()
    wb_new.remove(wb_new.active)
    print(f"--> Berkas '{final_output}' tidak ditemukan. Membuat berkas baru...")

df_ml = df_micro_master.copy()
df_ml.columns = [clean_key(c) for c in df_ml.columns]

features_cols = ["cleanednilaifaktur", "cleanedsisapiutang", "cleaneddays"]
X = df_ml[features_cols].fillna(0).values
scaler = StandardScaler()
X_scaled = scaler.fit_transform(X).astype(np.float32)
chunk_size = 20000

def format_sheet_table(ws, header_cols, data_matrix, number_formats=None):
    ws.append(header_cols)
    h_row = ws.max_row
    for c_idx in range(1, len(header_cols) + 1):
        cell = ws.cell(row=h_row, column=c_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
        
    for row_data in data_matrix:
        ws.append(row_data)
        curr_r = ws.max_row
        is_anomaly = row_data[-1] == "REVIU INDIKASI ANOMALI"
        for c_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=curr_r, column=c_idx)
            cell.border = thin_border
            
            if number_formats and c_idx in number_formats:
                cell.number_format = number_formats[c_idx]
                if "%" in number_formats[c_idx] or "#,##0" in number_formats[c_idx]:
                    cell.alignment = align_right
            elif isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0;(#,##0);"-"'
                cell.alignment = align_right
            if is_anomaly:
                cell.fill = fill_anomaly
                cell.font = font_bold

def generate_and_add_plot(df_plot, anomaly_col, method_name, ws, cell_loc):
    plt.figure(figsize=(7, 4.5))
    sns.set_theme(style="whitegrid")
    sns.scatterplot(
        data=df_plot,
        x="cleanednilaifaktur",
        y="cleanedsisapiutang",
        hue=anomaly_col,
        palette={"WAJAR": "#366092", "ANOMALI": "#C00000"},
        style=anomaly_col,
        markers={"WAJAR": "o", "ANOMALI": "X"},
        alpha=0.7,
        s=70
    )
    plt.title(f"Visualisasi Pencilan - {method_name}", fontsize=12, fontweight="bold", color="#1F497D")
    plt.xlabel("Nilai Faktur (IDR)")
    plt.ylabel("Sisa Piutang (IDR)")
    plt.tight_layout()
     
    img_filename = f"plot_{anomaly_col}.png"
    plt.savefig(img_filename, dpi=100)
    plt.close()
    
    img = Image(img_filename)
    ws.add_image(img, cell_loc)

def extract_anomaly_matrix(df_source, anomaly_col):
    df_anomali = df_source[df_source[anomaly_col] == "ANOMALI"]
    matrix = []
    for _, row in df_anomali.iterrows():
        no_faktur = str(row[no_faktur_col]) if no_faktur_col in row else "-"
        nama_cust = str(row[nama_pelanggan_col]) if nama_pelanggan_col in row else str(row.get('cleanedcustid', '-'))
        matrix.append([
            str(row.get('bulansistem', '-')),
            no_faktur,
            nama_cust,
            float(row.get('cleanednilaifaktur', 0)),
            float(row.get('cleanedsisapiutang', 0)),
            float(row.get('cleaneddays', 0)),
            "REVIU INDIKASI ANOMALI"
        ])
    if not matrix:
        matrix.append(["-", "-", "Tidak ada transaksi pencilan terdeteksi", 0, 0, 0, "WAJAR"])
    return matrix

headers_ml = ["Bulan", "Nomor Faktur", "ID/Nama Pelanggan", "Nilai Faktur", "Sisa Piutang", "Umur Piutang (Hari)", "Kesimpulan Audit"]
formats_ml = {4: '#,##0', 5: '#,##0', 6: '#,##0'}

print("--> Menyusun Sheet 6: Deteksi Anomali DBSCAN...")
if len(X_scaled) <= chunk_size:
    dbscan = DBSCAN(eps=0.5, min_samples=4, algorithm='ball_tree', leaf_size=100)
    labels_dbscan = dbscan.fit_predict(X_scaled)
else:
    labels_dbscan = np.zeros(len(X_scaled), dtype=int)
    for start_idx in range(0, len(X_scaled), chunk_size):
        end_idx = min(start_idx + chunk_size, len(X_scaled))
        X_chunk = X_scaled[start_idx:end_idx]
        dbscan_chunk = DBSCAN(eps=0.5, min_samples=4, algorithm='ball_tree', leaf_size=100)
        labels_dbscan[start_idx:end_idx] = dbscan_chunk.fit_predict(X_chunk)

df_ml["anomaly_dbscan"] = np.where(labels_dbscan == -1, "ANOMALI", "WAJAR")

ws6 = wb_new.create_sheet(title="6. Anomali ML DBSCAN")
ws6.append(["6. PROSEDUR AUDIT ADVANCED: DETEKSI PENCIAN BERBASIS DENSITAS (DBSCAN)"])
ws6.cell(row=1, column=1).font = font_title
ws6.append(["Fokus Audit: Menemukan transaksi yang terisolasi di area berkerapatan rendah (pencilan ekstrem struktural)."])
ws6.append([])

matrix_s6 = extract_anomaly_matrix(df_ml, "anomaly_dbscan")
format_sheet_table(ws6, headers_ml, matrix_s6, formats_ml)
generate_and_add_plot(df_ml, "anomaly_dbscan", "DBSCAN Clustering", ws6, "I5")

print("--> Menyusun Sheet 7: Deteksi Anomali Isolation Forest...")
iso_forest = IsolationForest(contamination=0.02, random_state=42, n_jobs=-1)
labels_iso = iso_forest.fit_predict(X_scaled)
df_ml["anomaly_iso"] = np.where(labels_iso == -1, "ANOMALI", "WAJAR")

ws7 = wb_new.create_sheet(title="7. Anomali Isolation Forest")
ws7.append(["7. PROSEDUR AUDIT ADVANCED: ANOMALI STRUKTURAL (ISOLATION FOREST)"])
ws7.cell(row=1, column=1).font = font_title
ws7.append(["Fokus Audit: Mengisolasi anomali menggunakan struktur pohon biner. Transaksi yang cepat terisolasi dianggap berisiko tinggi."])
ws7.append([])

matrix_s7 = extract_anomaly_matrix(df_ml, "anomaly_iso")
format_sheet_table(ws7, headers_ml, matrix_s7, formats_ml)
generate_and_add_plot(df_ml, "anomaly_iso", "Isolation Forest", ws7, "I5")

print("--> Menyusun Sheet 8: Deteksi Anomali Local Outlier Factor...")
if len(X_scaled) <= chunk_size:
    lof = LocalOutlierFactor(n_neighbors=20, contamination=0.02, algorithm='ball_tree', leaf_size=100)
    labels_lof = lof.fit_predict(X_scaled)
else:
    labels_lof = np.ones(len(X_scaled), dtype=int)
    for start_idx in range(0, len(X_scaled), chunk_size):
        end_idx = min(start_idx + chunk_size, len(X_scaled))
        X_chunk = X_scaled[start_idx:end_idx]
        lof_chunk = LocalOutlierFactor(n_neighbors=20, contamination=0.02, algorithm='ball_tree', leaf_size=100)
        labels_lof[start_idx:end_idx] = lof_chunk.fit_predict(X_chunk)

df_ml["anomaly_lof"] = np.where(labels_lof == -1, "ANOMALI", "WAJAR")

ws8 = wb_new.create_sheet(title="8. Anomali LOF Local")
ws8.append(["8. PROSEDUR AUDIT ADVANCED: KEPADATAN LOKAL (LOCAL OUTLIER FACTOR)"])
ws8.cell(row=1, column=1).font = font_title
ws8.append(["Fokus Audit: Membandingkan densitas lokal suatu transaksi dengan tetangganya untuk melihat penyimpangan perilaku lokal."])
ws8.append([])

matrix_s8 = extract_anomaly_matrix(df_ml, "anomaly_lof")
format_sheet_table(ws8, headers_ml, matrix_s8, formats_ml)
generate_and_add_plot(df_ml, "anomaly_lof", "Local Outlier Factor", ws8, "I5")

print("--> Menyusun Sheet 9: Deteksi Anomali One-Class SVM...")
oc_svm = OneClassSVM(nu=0.02, kernel="rbf", gamma="scale")
if len(X_scaled) > 10000:
    np.random.seed(42)
    sample_idx = np.random.choice(len(X_scaled), 10000, replace=False)
    oc_svm.fit(X_scaled[sample_idx])
else:
    oc_svm.fit(X_scaled)

labels_svm = oc_svm.predict(X_scaled)
df_ml["anomaly_svm"] = np.where(labels_svm == -1, "ANOMALI", "WAJAR")

ws9 = wb_new.create_sheet(title="9. Anomali One-Class SVM")
ws9.append(["9. PROSEDUR AUDIT ADVANCED: BATAS KEPUTUSAN KERNEL NON-LINEAR (ONE-CLASS SVM)"])
ws9.cell(row=1, column=1).font = font_title
ws9.append(["Fokus Audit: Memetakan pola transaksi normal ke ruang berdimensi tinggi untuk memisahkan pencilan secara non-linear."])
ws9.append([])

matrix_s9 = extract_anomaly_matrix(df_ml, "anomaly_svm")
format_sheet_table(ws9, headers_ml, matrix_s9, formats_ml)
generate_and_add_plot(df_ml, "anomaly_svm", "One-Class SVM", ws9, "I5")

print("--> Menyusun Sheet 10: Deteksi Anomali K-Means Distance...")
kmeans = KMeans(n_clusters=4, random_state=42, n_init=3)
kmeans.fit(X_scaled)
centroids = kmeans.cluster_centers_
km_labels = kmeans.labels_
distances = np.sqrt(np.sum((X_scaled - centroids[km_labels])**2, axis=1))
threshold_km = np.percentile(distances, 98)
df_ml["anomaly_kmeans"] = np.where(distances > threshold_km, "ANOMALI", "WAJAR")

ws10 = wb_new.create_sheet(title="10. Anomali K-Means Cluster")
ws10.append(["10. PROSEDUR AUDIT ADVANCED: DISTANCE-BASED CENTROID (K-MEANS CLUSTERING)"])
ws10.cell(row=1, column=1).font = font_title
ws10.append(["Fokus Audit: Menemukan transaksi dengan jarak geometris terjauh dari pusat kelompok (centroid) normalnya."])
ws10.append([])

matrix_s10 = extract_anomaly_matrix(df_ml, "anomaly_kmeans")
format_sheet_table(ws10, headers_ml, matrix_s10, formats_ml)
generate_and_add_plot(df_ml, "anomaly_kmeans", "K-Means Centroid Distance", ws10, "I5")

print("--> Menyusun Sheet 11: Deteksi Anomali Neural Network Autoencoder...")
autoencoder = MLPRegressor(hidden_layer_sizes=(2,), activation="relu", max_iter=100, early_stopping=True, random_state=42)
autoencoder.fit(X_scaled, X_scaled)
X_reconstructed = autoencoder.predict(X_scaled)
recon_error = np.mean((X_scaled - X_reconstructed)**2, axis=1)
threshold_ae = np.percentile(recon_error, 98)
df_ml["anomaly_ae"] = np.where(recon_error > threshold_ae, "ANOMALI", "WAJAR")

ws11 = wb_new.create_sheet(title="11. Anomali Autoencoder NN")
ws11.append(["11. PROSEDUR AUDIT ADVANCED: RECONSTRUCTION ERROR (NEURAL NETWORK AUTOENCODER)"])
ws11.cell(row=1, column=1).font = font_title
ws11.append(["Fokus Audit: Mengidentifikasi transaksi dengan error rekonstruksi tertinggi (pola fraud/salah input kompleks non-linear)."])
ws11.append([])

matrix_s11 = extract_anomaly_matrix(df_ml, "anomaly_ae")
format_sheet_table(ws11, headers_ml, matrix_s11, formats_ml)
generate_and_add_plot(df_ml, "anomaly_ae", "Neural Network Autoencoder", ws11, "I5")

print("--> Menyelaraskan ukuran lebar kolom seluruh berkas audit otomatis...")
for sheet_name in wb_new.sheetnames:
    ws_to_fit = wb_new[sheet_name]
    for col in ws_to_fit.columns:
        col_letter = get_column_letter(col[0].column)
        
        if col_letter == 'A':
            ws_to_fit.column_dimensions[col_letter].width = 35
        else:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws_to_fit.column_dimensions[col_letter].width = max(max_len + 3, 12)

wb_new.save(final_output)

print("--> Membersihkan berkas gambar temporary di dalam folder...")
list_algoritma = [
    "ae", 
    "dbscan", 
    "iso", 
    "kmeans", 
    "lof", 
    "svm"
]

print("--> Memulai pembersihan file sementara...")
for algo in list_algoritma:
    nama_file = f"plot_anomaly_{algo.replace(' ', '_').lower()}.png"
    if os.path.exists(nama_file):
            os.remove(nama_file)

print(f"--> SUKSES! Berkas laporan analitis audit '{final_output}' telah diperbarui dengan prosedur Machine Learning.")
