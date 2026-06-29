import openpyxl
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font

print("--> Memulai proses penyuntikkan diagram dinamis...")

output_file = "Monitoring.xlsx"
wb = openpyxl.load_workbook(output_file)
ws = wb["Rekap Monitoring"]

header_row_idx = None
for r in range(1, ws.max_row + 1):
    if ws.cell(row=r, column=1).value == "KETERANGAN":
        header_row_idx = r
        break

if not header_row_idx:
    print("--> Error: Struktur data tidak dikenali. Baris KETERANGAN tidak ditemukan.")
    exit()

months_cols = []
for c in range(3, ws.max_column + 1):
    val = ws.cell(row=header_row_idx, column=c).value
    if val == "" or val is None:
        break
    months_cols.append(c)

min_month_col = months_cols[0]
max_month_col = months_cols[-1]

row_total_ar = None
row_ar_60 = None
row_bad_debt = None

for r in range(header_row_idx + 1, ws.max_row + 1):
    val = str(ws.cell(row=r, column=1).value).strip()
    if val == "TOTAL AR OUTSTANDING (AMOUNT)":
        row_total_ar = r
    elif val == "AR 60 HARI UP (AMOUNT)":
        row_ar_60 = r
    elif val == "AR BAD DEBT (365 UP ) AMOUNT":
        row_bad_debt = r

bridge_title_row = ws.max_row + 3
ws.cell(row=bridge_title_row, column=1, value="[ TABEL ACUAN DIAGRAM TREN BULANAN ]").font = Font(name="Calibri", size=10, italic=True, color="595959")

bridge_start_row = bridge_title_row + 1

ws.cell(row=bridge_start_row, column=1, value="Keterangan")
for idx, c_idx in enumerate(months_cols):
    m_name = ws.cell(row=header_row_idx, column=c_idx).value
    ws.cell(row=bridge_start_row, column=2 + idx, value=m_name)

ws.cell(row=bridge_start_row + 1, column=1, value="TOTAL AR OUTSTANDING (AMOUNT)")
for idx, c_idx in enumerate(months_cols):
    val = ws.cell(row=row_total_ar, column=c_idx).value
    ws.cell(row=bridge_start_row + 1, column=2 + idx, value=val)

ws.cell(row=bridge_start_row + 2, column=1, value="AR 60 HARI UP (AMOUNT)")
for idx, c_idx in enumerate(months_cols):
    val = ws.cell(row=row_ar_60, column=c_idx).value
    ws.cell(row=bridge_start_row + 2, column=2 + idx, value=val)

ws.cell(row=bridge_start_row + 3, column=1, value="AR BAD DEBT (365 UP ) AMOUNT")
for idx, c_idx in enumerate(months_cols):
    val = ws.cell(row=row_bad_debt, column=c_idx).value
    ws.cell(row=bridge_start_row + 3, column=2 + idx, value=val)

chart_bar = BarChart()
chart_bar.type = "col"
chart_bar.style = 10
chart_bar.title = "Tren AR Outstanding & Umur Piutang Bulanan"
chart_bar.y_axis.title = "Jumlah (Amount)"
chart_bar.x_axis.title = "Bulan"
chart_bar.width = 18
chart_bar.height = 10

max_bridge_col = 1 + len(months_cols)
data_ref = Reference(ws, min_col=1, min_row=bridge_start_row + 1, max_col=max_bridge_col, max_row=bridge_start_row + 3)
cats_ref = Reference(ws, min_col=2, min_row=bridge_start_row, max_col=max_bridge_col, max_row=bridge_start_row)

chart_bar.add_data(data_ref, titles_from_data=True, from_rows=True)
chart_bar.set_categories(cats_ref)

start_r = None
end_r = None
for r in range(header_row_idx + 1, ws.max_row + 1):
    val = str(ws.cell(row=r, column=1).value).strip()
    if "Daftar Kontribusi Barang AR-60 HARI" in val:
        start_r = r + 1
    elif val == "Total" and start_r is not None:
        end_r = r - 1
        break

chart_pie = PieChart()
latest_month_name = ws.cell(row=header_row_idx, column=max_month_col).value
chart_pie.title = f"Kontribusi Barang pada AR 60 Hari Up ({latest_month_name})"
chart_pie.width = 14
chart_pie.height = 10

if start_r and end_r:
    labels = Reference(ws, min_col=1, min_row=start_r, max_row=end_r)
    data_pie = Reference(ws, min_col=max_month_col, min_row=start_r, max_row=end_r)
    chart_pie.add_data(data_pie, titles_from_data=False)
    chart_pie.set_categories(labels)

start_chart_row = ws.max_row + 3
ws.add_chart(chart_bar, f"A{start_chart_row}")
if start_r and end_r:
    ws.add_chart(chart_pie, f"L{start_chart_row}")

start_text_row = start_chart_row + 21
font_title = Font(name="Calibri", size=11, bold=True)
font_desc = Font(name="Calibri", size=11, italic=True)

ws.cell(row=start_text_row, column=1, value="KETERANGAN DAN ANALISIS DIAGRAM:").font = font_title
ws.cell(row=start_text_row + 1, column=1, value="1. Diagram Batang (Tren Bulanan):").font = font_title
ws.cell(row=start_text_row + 2, column=1, value="Menampilkan fluktuasi Total AR Outstanding berdampingan dengan penuaan piutang (60 Hari Up & Bad Debt). Berguna untuk memantau apakah lonjakan outstanding bulanan diimbangi dengan kontrol penagihan yang ketat atau justru memperbanyak piutang macet.").font = font_desc

ws.cell(row=start_text_row + 4, column=1, value="2. Diagram Kue (Kontribusi Produk):").font = font_title
ws.cell(row=start_text_row + 5, column=1, value=f"Menggambarkan persentase pembentuk saldo AR 60 Hari Up berdasarkan kelompok filter produk pada periode terbaru ({latest_month_name}). Membantu manajemen mengidentifikasi produk utama yang menyumbang risiko kredit terbesar.").font = font_desc

wb.save(output_file)
print(f"--> Selesai sempurna! Diagram dan keterangan berhasil disuntikkan ke '{output_file}'.")